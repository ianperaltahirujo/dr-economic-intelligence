"""
DR Economic Vulnerability Pipeline - Excel Output Writer
Produces a professionally formatted Excel workbook from pipeline results.

Sheets:
    Dashboard   - Single-page summary: score, status, top indicators
    Contexto    - Non-scored contextual macro indicators (Tourism, Debt)
    Indicators  - All 12 vulnerability components with z-scores and weights
    Alerts      - Plain-language flagged indicators for Claude briefings
    History     - Last 60 months of scores and raw indicator values
    Metadata    - Pipeline run info, methodology, data source freshness

Primary reader: Claude via Microsoft 365 MCP integration
Secondary reader: Human stakeholders at La Sociedad

Usage:
    from pipeline.write_excel import write_workbook
    write_workbook(results, path="data/output/vulnerability_report.xlsx")
"""

import os
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))

from pipeline.build_vulnerability import (
    VULNERABILITY_COMPONENTS,
    INDICATOR_LABELS,
    HIGH_STRESS_THRESHOLD,
    classify_indicator
)

NAVY       = "0D1B2A"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F6FA"
MID_GRAY   = "D1D5DB"
DARK_GRAY  = "6B7280"
STRESS_RED = "DC2626"
WATCH_ORG  = "D97706"
GOOD_GRN   = "16A34A"

INDICATOR_FORMATS = {
    "remesas_usd_mm":           "#,##0",
    "ipc_yoy_pct":              "0.00%",
    "dop_usd":                  "0.00",
    "reserves_usd_mm":          "#,##0",
    "imae_index":               "0.0",
    "sb_morosidad_pct":         "0.00%",
    "sb_solvencia_pct":         "0.00%",
    "UNRATE":                   "0.00%",
    "UMCSENT":                  "0.0",
    "sb_tasa_activa_pct":       "0.00%",
    "gas_premium_dop":          "#,##0.00",
    "tourism_daily_spend_usd":  "#,##0.00",
}

def _set_border(cell) -> None:
    thin = Side(border_style="thin", color=MID_GRAY)
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def _style_header(cell, bg_color=NAVY, fg_color=WHITE) -> None:
    cell.font = Font(bold=True, color=fg_color)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    _set_border(cell)

def _build_dashboard(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    score      = results.get("current_score", 0) or 0
    score_date = results.get("score_date")
    scored     = results.get("scored", pd.DataFrame())
    
    is_provisional = False
    if score_date and "is_provisional" in scored.columns and score_date in scored.index:
        is_provisional = bool(scored.loc[score_date, "is_provisional"])
    
    if score >= HIGH_STRESS_THRESHOLD: status_text, status_color = "ESTRÉS ALTO", STRESS_RED
    elif score >= 50: status_text, status_color = "ESTRÉS MODERADO", WATCH_ORG
    else: status_text, status_color = "NORMAL", GOOD_GRN

    ws.cell(row=2, column=2, value="DR ECONOMIC VULNERABILITY INDEX").font = Font(size=18, bold=True, color=NAVY)
    label_font, val_font = Font(size=12, color=DARK_GRAY, bold=True), Font(size=14, bold=True, color=NAVY)

    ws.cell(row=4, column=2, value="Score:").font = label_font
    score_cell = ws.cell(row=4, column=3, value=round(score, 1))
    score_cell.font = Font(size=18, bold=True, color=status_color)
    score_cell.alignment = Alignment(horizontal="left")

    if score_date:
        ws.cell(row=5, column=2, value="Fecha:").font = label_font
        date_display = score_date.strftime("%B %Y").capitalize()
        if is_provisional: date_display += " (Avance Estimado)"
        ws.cell(row=5, column=3, value=date_display).font = val_font

    ws.cell(row=6, column=2, value="Status:").font = label_font
    ws.cell(row=6, column=3, value=status_text).font = Font(size=14, bold=True, color=status_color)

    row = 9
    n_total = len(VULNERABILITY_COMPONENTS)
    ws.cell(row=row, column=2, value=f"Componentes Clave ({n_total}/{n_total} requeridos)").font = Font(size=14, bold=True, color=NAVY)
    row += 2

    for c_idx, header in enumerate(["Indicador", "Valor", "Tendencia", "Estado", "Z-Score"], start=2):
        _style_header(ws.cell(row=row, column=c_idx))
    row += 1

    for col, (weight, direction) in VULNERABILITY_COMPONENTS.items():
        zscore_col = f"{col}_zscore"
        if score_date not in scored.index or col not in scored.columns or zscore_col not in scored.columns: continue
        
        value = scored.loc[score_date, col]
        zscore = scored.loc[score_date, zscore_col]
        if pd.isna(value) or pd.isna(zscore): continue

        prev_idx = scored.index[scored.index < score_date]
        delta = value - scored.loc[prev_idx[-1], col] if not prev_idx.empty else 0

        classification = classify_indicator(col, value, zscore)
        if classification["is_stress"]: status, bg_color, text_color = "ALERTA", "FEE2E2", "991B1B"
        elif classification["is_watch"]: status, bg_color, text_color = "VIGILANCIA", "DBEAFE", "1E40AF"
        else: status, bg_color, text_color = "NORMAL", "F3F4F6", "374151"

        if delta == 0: trend_arrow, trend_color = "-", DARK_GRAY
        elif direction == "positive": trend_arrow, trend_color = ("↑", STRESS_RED) if delta > 0 else ("↓", GOOD_GRN)
        else: trend_arrow, trend_color = ("↑", GOOD_GRN) if delta > 0 else ("↓", STRESS_RED)

        c_name = ws.cell(row=row, column=2, value=INDICATOR_LABELS.get(col, col))
        c_name.font = Font(bold=True, color=NAVY)
        _set_border(c_name)
        
        fmt_val = value / 100.0 if col in ["ipc_yoy_pct", "sb_morosidad_pct", "sb_solvencia_pct", "UNRATE", "sb_tasa_activa_pct"] else value
        c_val = ws.cell(row=row, column=3, value=fmt_val)
        c_val.number_format = INDICATOR_FORMATS.get(col, "0.00")
        c_val.alignment = Alignment(horizontal="right")
        _set_border(c_val)
        
        c_trend = ws.cell(row=row, column=4, value=trend_arrow)
        c_trend.font = Font(bold=True, color=trend_color)
        c_trend.alignment = Alignment(horizontal="center")
        _set_border(c_trend)
        
        c_stat = ws.cell(row=row, column=5, value=status)
        c_stat.font = Font(bold=True, color=text_color)
        c_stat.fill = PatternFill("solid", fgColor=bg_color)
        c_stat.alignment = Alignment(horizontal="center")
        _set_border(c_stat)
        
        c_z = ws.cell(row=row, column=6, value=zscore)
        c_z.number_format = "+0.00;-0.00;0.00"
        c_z.alignment = Alignment(horizontal="right")
        _set_border(c_z)
        row += 1

    ws.column_dimensions["A"].width, ws.column_dimensions["B"].width, ws.column_dimensions["C"].width, ws.column_dimensions["D"].width, ws.column_dimensions["E"].width, ws.column_dimensions["F"].width = 3, 40, 15, 12, 18, 15

def _build_context(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Contexto")
    ws.sheet_view.showGridLines = False
    row = 2
    ws.cell(row=row, column=2, value="Contexto Macroeconómico (No Puntuado)").font = Font(size=16, bold=True, color=NAVY)
    row += 2

    datasets = [
        ("Sector Turismo - Recaudación Fiscal (Millones DOP)", results.get("tourism_fiscal", pd.DataFrame())),
        ("Deuda Pública Consolidada (Millones USD / % PIB)", results.get("debt_detail", pd.DataFrame()))
    ]

    for title, df in datasets:
        if df.empty: continue
        ws.cell(row=row, column=2, value=title).font = Font(size=12, bold=True, color=NAVY)
        row += 1
        for c_idx, col_name in enumerate(["Fecha"] + list(df.columns), start=2): _style_header(ws.cell(row=row, column=c_idx), bg_color=NAVY)
        row += 1
        for date_idx, data_row in df.tail(24)[::-1].iterrows():
            _set_border(ws.cell(row=row, column=2, value=date_idx.strftime("%Y-%m-%d")))
            for c_idx, col_name in enumerate(df.columns, start=3):
                c_val = ws.cell(row=row, column=c_idx, value=data_row[col_name] if pd.notna(data_row[col_name]) else "")
                c_val.number_format = "0.00%" if "pct" in col_name.lower() else "#,##0.00"
                _set_border(c_val)
            row += 1
        row += 2

    ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 3, 15
    for col_letter in ["C", "D", "E", "F", "G", "H"]: ws.column_dimensions[col_letter].width = 25

def _build_indicators(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Indicators")
    scored, score_date = results.get("scored", pd.DataFrame()), results.get("score_date")
    for c_idx, header in enumerate(["ID", "Indicador", "Valor Reciente", "Fecha", "Z-Score", "Peso", "Impacto Bruto"], start=1): _style_header(ws.cell(row=1, column=c_idx))

    row = 2
    for col, (weight, direction) in VULNERABILITY_COMPONENTS.items():
        if score_date not in scored.index or col not in scored.columns: continue
        val, zscore = scored.loc[score_date, col], scored.loc[score_date, f"{col}_zscore"]
        if pd.isna(val): continue
        ws.cell(row=row, column=1, value=col)
        ws.cell(row=row, column=2, value=INDICATOR_LABELS.get(col, col))
        ws.cell(row=row, column=3, value=val)
        ws.cell(row=row, column=4, value=score_date.strftime("%Y-%m-%d"))
        ws.cell(row=row, column=5, value=zscore).number_format = "0.00"
        ws.cell(row=row, column=6, value=weight).number_format = "0%"
        ws.cell(row=row, column=7, value=(zscore * weight if direction == "positive" else -zscore * weight)).number_format = "0.000"
        row += 1
    for c in ["A", "B", "C", "D", "E", "F", "G"]: ws.column_dimensions[c].width = 20
    ws.column_dimensions["B"].width = 40

def _build_alerts(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Alerts")
    alerts = results.get("alerts", pd.DataFrame())
    for c_idx, header in enumerate(["Indicador", "Nivel", "Análisis"], start=1): _style_header(ws.cell(row=1, column=c_idx), bg_color=STRESS_RED)

    if alerts.empty: ws.cell(row=2, column=1, value="No active alerts.")
    else:
        for row, (_, alert) in enumerate(alerts.iterrows(), start=2):
            ws.cell(row=row, column=1, value=INDICATOR_LABELS.get(alert["indicator"], alert["indicator"]))
            ws.cell(row=row, column=2, value="ESTRÉS" if alert.get("is_stress") else "VIGILANCIA").font = Font(bold=True, color=STRESS_RED if alert.get("is_stress") else WATCH_ORG)
            ws.cell(row=row, column=3, value=alert.get("alert_text", ""))
    ws.column_dimensions["A"].width, ws.column_dimensions["B"].width, ws.column_dimensions["C"].width = 30, 15, 120

def _build_history(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("History")
    scored = results.get("scored", pd.DataFrame())
    if scored.empty: return
    
    cols = ["vulnerability_score", "is_provisional"] + [c for c in VULNERABILITY_COMPONENTS.keys() if c in scored.columns]
    
    _style_header(ws.cell(row=1, column=1))
    ws.cell(row=1, column=1, value="Date")
    for c_idx, col in enumerate(cols, start=2):
        _style_header(ws.cell(row=1, column=c_idx))
        ws.cell(row=1, column=c_idx, value=col)

    for row, (date_idx, data_row) in enumerate(scored.tail(60)[::-1].iterrows(), start=2):
        ws.cell(row=row, column=1, value=date_idx.strftime("%Y-%m-%d"))
        for c_idx, col in enumerate(cols, start=2):
            val = data_row[col] if col in data_row else ""
            if pd.notna(val): ws.cell(row=row, column=c_idx, value=val)

    ws.column_dimensions["A"].width = 15
    for i in range(2, len(cols) + 2): ws.column_dimensions[get_column_letter(i)].width = 20

def _build_metadata(wb: Workbook, results: dict) -> None:
    ws = wb.create_sheet("Metadata")
    ws.column_dimensions["A"].width, ws.column_dimensions["B"].width = 3, 100
    r = 2
    ws.cell(row=r, column=2, value="System Metadata & Methodology").font = Font(size=14, bold=True)
    r += 2
    ws.cell(row=r, column=2, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    r += 1
    sd_str = results.get("score_date").strftime('%Y-%m-%d') if results.get("score_date") else "None"
    ws.cell(row=r, column=2, value=f"Headline Score Date: {sd_str}")
    r += 2
    ws.cell(row=r, column=2, value="Methodology Rules (Engine v4):").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=2, value="• Score Coverage: Exige 12 de 12 indicadores. Sin embargo, realiza un *nowcast* proyectando hasta 2 meses hacia el futuro los datos rezagados para proveer reportes semanales oportunos. Cualquier mes que contenga proyecciones es marcado estrictamente como 'is_provisional = True'.")
    r += 1
    ws.cell(row=r, column=2, value="• Z-Score Window: 36 meses móviles (60 meses para IPC y Tasa de Cambio). Exige un mínimo de 12 meses históricos.")
    r += 1
    ws.cell(row=r, column=2, value="• Alert Thresholds: Flag de VIGILANCIA en |z| > 0.75 y ESTRÉS en |z| > 1.5 en dirección de riesgo económico.")
    r += 1
    ws.cell(row=r, column=2, value="• Absolute Level Overrides: Rupturas de umbrales absolutos fuerzan automáticamente el estatus a ESTRÉS.")

def write_workbook(results: dict, path: str = "data/output/vulnerability_report.xlsx") -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    print("Writing Excel workbook...")
    _build_dashboard(wb, results); print("  [OK] Dashboard")
    _build_context(wb, results); print("  [OK] Contexto")
    _build_indicators(wb, results); print("  [OK] Indicators")
    _build_alerts(wb, results); print("  [OK] Alerts")
    _build_history(wb, results); print("  [OK] History")
    _build_metadata(wb, results); print("  [OK] Metadata")

    wb.save(output_path)
    print(f"\nWorkbook saved: {output_path}")
    return output_path

if __name__ == "__main__":
    from pipeline.build_vulnerability import run_vulnerability_pipeline
    print("Running vulnerability pipeline...")
    res = run_vulnerability_pipeline()
    write_workbook(res)